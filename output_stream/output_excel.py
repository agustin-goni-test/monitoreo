from .output_writer import OutputWriter
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from polling.poller import TransactionPolling
from datetime import datetime
import openpyxl
import glob
import re

class ExcelWriter(OutputWriter):
    def write_default(self, service_name: str, data_matrix: list[list], **kwargs):
        sheet_name = kwargs.get('sheet_name', 'Sheet1')
        
        output_dir = "output_files"
        os.makedirs(output_dir, exist_ok=True)
        
        base_filename = f"{service_name}_output.xlsx"
        filename = os.path.join(output_dir, base_filename)
        counter = 1
        while os.path.exists(filename):
            filename = os.path.join(output_dir, f"{service_name}_output_{counter}.xlsx")
            counter += 1
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write header
        header = data_matrix[0]
        ws.append(header)
        
        # Infer column formats once
        col_formats = []
        for col_name in header:
            col_name_lower = col_name.lower()
            if "timestamp" in col_name_lower:
                col_formats.append("text")  # You can refine this if needed
            elif "time" in col_name_lower:
                col_formats.append("time")
            elif "rate" in col_name_lower or "%" in col_name_lower:
                col_formats.append("percentage")
            elif "count" in col_name_lower:
                col_formats.append("integer")
            else:
                col_formats.append("text")
        
        # Write data rows
        for row in data_matrix[1:]:
            output_row = []
            for value, fmt in zip(row, col_formats):
                if value is None:
                    output_row.append("")
                else:
                    try:
                        if fmt == "integer":
                            output_row.append(int(value))
                        elif fmt == "time":
                            output_row.append(float(value))  # Assuming time in seconds
                        elif fmt == "percentage":
                            output_row.append(float(value))  # Assuming decimal form like 0.05 = 5%
                        else:
                            output_row.append(str(value))
                    except (ValueError, TypeError):
                        output_row.append(str(value))
            ws.append(output_row)
        
        # Apply formats
        for col_idx, fmt in enumerate(col_formats, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for cell in ws[col_letter][1:]:  # Skip header
                if fmt == "percentage":
                    cell.number_format = '0.00%'
                elif fmt == "time":
                    cell.number_format = '0.00'
                elif fmt == "integer":
                    cell.number_format = '0'
                # Timestamps and text: leave as is
        
        wb.save(filename)
        
        print(f"\nWrote a total of {len(data_matrix) - 1} data points, with {len(header) - 1} columns, to the Excel file '{filename}' (sheet: '{sheet_name}')")


    def write_last_trx_poll(self, polling_data: TransactionPolling):
        """
        Write transaction polling data to a daily Excel file.
        """
        output_dir = "output_files"
        os.makedirs(output_dir, exist_ok=True)

        date_str = datetime.now().strftime("%Y_%m_%d")
        base_filename = f"trx_polling_{date_str}.xlsx"

        # Check for an existing file ignoring any timestamped versions
        existing_file = self._find_existing_file(output_dir, date_str)

        if existing_file:
            filename = os.path.join(output_dir, existing_file)
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
        else:
            filename = os.path.join(output_dir, base_filename)
            workbook = Workbook()
            sheet = workbook.active

            # Write header row
            sheet.append([
                "Poll time",
                "Year",
                "Month",
                "Day",
                "Weekday",
                "Time",
                "Last transaction",
                "lag (min)",
                "Lag minutes",
                "Lag seconds"
            ])
            
        # Extract values
        poll_time = polling_data.current_time
        last_trx_time = polling_data.last_transaction_time
        lag_str = f"{polling_data.time_lag:.2f}"
        lag = float(lag_str)

        # Decomposed poll time
        year = poll_time.year
        month = poll_time.month
        day = poll_time.day
        weekday = poll_time.strftime("%A")
        time_str = poll_time.strftime("%H:%M:%S")

        total_seconds = lag * 60
        lag_minutes = int(total_seconds // 60)
        lag_seconds = total_seconds % 60

        # Create new row
        row = [
            poll_time.isoformat(),
            year,
            month,
            day,
            weekday,
            time_str,
            last_trx_time.isoformat(),
            lag,
            lag_minutes,
            lag_seconds
        ]   

        # Write the data row
        sheet.append(row)

        workbook.save(filename)
        print(f"Appended polling data to {filename}")


    def finalize_last_trx_poll_file(self):
        """
        Rename the current day's polling file to include a timestamp when the process is interrupted.
        """
        output_dir = "output_files"
        date_str = datetime.now().strftime("%Y_%m_%d")

        # Find the file ignoring timestamp suffix
        existing_file = self._find_existing_file(output_dir, date_str)
        if existing_file:
            old_path = os.path.join(output_dir, existing_file)
            base_name, ext = os.path.splitext(existing_file)

            # Remove any existing timestamp suffix
            if " - " in base_name:
                base_name = base_name.split(" - ")[0]

            timestamp_str = datetime.now().strftime("%H_%M_%S")
            new_name = f"{base_name} - {timestamp_str}{ext}"
            new_path = os.path.join(output_dir, new_name)

            os.rename(old_path, new_path)
            print(f"Renamed file to: {new_name}")


    def write_polling_stats(self, service_name, stats_list):
        """
        Write polling stats to an Excel file per service.
        Each metric is written to a separate sheet named after the metric.
        """
        output_dir = "output_files"
        os.makedirs(output_dir, exist_ok=True)

        base_filename = f"Poll_{service_name}"
        pattern = os.path.join(output_dir, f"{base_filename}*.xlsx")
        matches = glob.glob(pattern)

        if matches:
            filename = matches[0]
            workbook = openpyxl.load_workbook(filename)
        else:
            filename = os.path.join(output_dir, f"{base_filename}.xlsx")
            workbook = Workbook()
            # Remove the default sheet if it’s empty
            default_sheet = workbook.active
            if default_sheet and default_sheet.max_row == 1 and default_sheet.max_column == 1 and default_sheet.cell(1, 1).value is None:
                workbook.remove(default_sheet)

        # Define header row (without Metric Name, as each sheet is specific)
        header = [
            "Poll time",
            "Mean (ms)",
            "Median (ms)",
            "Min (ms)",
            "Max (ms)",
            "StdDev (ms)",
            "Compliance"
        ]

        # Define max sheet name length (Excel allows 31 chars max)
        MAX_SHEET_NAME_LENGTH = 31

        def truncate_sheet_name(name):
            """Truncate the sheet name consistently."""
            return name[:MAX_SHEET_NAME_LENGTH]

        poll_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for metric_name, stats in stats_list:
            sheet_name = truncate_sheet_name(metric_name)

            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(title=sheet_name)
                sheet.append(header)

            row = [
                poll_time,
                f"{stats.mean:.2f}",
                f"{stats.median:.2f}",
                f"{stats.min:.2f}",
                f"{stats.max:.2f}",
                f"{stats.std_dev:.2f}",
                str(stats.compliance)
            ]
            sheet.append(row)

        workbook.save(filename)
        print(f"Appended polling data to {filename}")


    def finalize_polling_file(self, service_name):
        """
        Rename the polling file to include a timestamp (to the second),
        ignoring files that already have a timestamp.
        """
        output_dir = "output_files"
        base_filename = f"Poll_{service_name}"
        pattern = os.path.join(output_dir, f"{base_filename}*.xlsx")
        matches = glob.glob(pattern)

        if matches:
            # Use the first match (you could refine this to pick the latest if needed)
            filename = matches[0]
            base_name, ext = os.path.splitext(os.path.basename(filename))
            # Remove any existing timestamp suffix at the end
            base_name = re.sub(r' - \d{2}-\d{2}-\d{2}$', '', base_name)
            timestamp_str = datetime.now().strftime("%H-%M-%S")
            new_filename = f"{base_name} - {timestamp_str}{ext}"
            new_filepath = os.path.join(output_dir, new_filename)
            os.rename(filename, new_filepath)
            print(f"Renamed file to: {new_filepath}")
        else:
            print(f"No file found to finalize for service '{service_name}'.")



    
    # Helper method to handle file names
    def _find_existing_file(self, directory, date_str):
        """
        Helper to find the file for today's date ignoring any timestamp suffix.
        """
        for filename in os.listdir(directory):
            if filename.startswith(f"trx_polling_{date_str}") and filename.endswith(".xlsx"):
                return filename
        return None
