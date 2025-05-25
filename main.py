import requests
import datetime
import pandas as pd
from dotenv import load_dotenv
import os
import time
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

# --- Configuration ---


load_dotenv()
API_TOKEN = os.getenv("DYNATRACE_API_TOKEN")

ENV_ID = 'kae68552'
OUTPUT_FOLDER = "output/"

# Service name → ID mapping
SERVICE_MAP = {
    "ComercioTransaccionesController": "SERVICE-FD9343224D905203",
    "AbonosController": "SERVICE-123E236BA4855F4A"
}


# --- Functions ---


def get_service_performance(service_id, metric_selector):
    """
    Fetch metrics for a given service from Dynatrace.
    """
    url = f'https://{ENV_ID}.live.dynatrace.com/api/v2/metrics/query'
    headers = {
        'Authorization': f'Api-Token {API_TOKEN}'
    }
    params = {
        'metricSelector': metric_selector,
        'resolution': '1m',
        'entitySelector': f'entityId({service_id})',
        'from': 'now-2d',
        'to': 'now'
    }
    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()
    return response.json()

def get_all_service_metrics(service_id):
    return {
        "response_time": get_service_performance(service_id, 'builtin:service.response.time'),
        "request_count": get_service_performance(service_id, 'builtin:service.requestCount.total'),
        "success_count": get_service_performance(service_id, 'builtin:service.errors.server.successCount'),
        "failure_count": get_service_performance(service_id, 'builtin:service.errors.server.count'),
    }


def parse_metric_data(data):
    series = data['result'][0]['data'][0]
    timestamps = series['timestamps']
    values = series['values']
    return {ts: val for ts, val in zip(timestamps, values)}


def poll_response_time(service_name, service_id, threshold):
    """
    Fetch average response time over the last 5 minutes and compare against threshold.
    """
    # Query last 5 minutes data from Dynatrace
    url = f'https://{ENV_ID}.live.dynatrace.com/api/v2/metrics/query'
    headers = {
        'Authorization': f'Api-Token {API_TOKEN}'
    }
    params = {
        'metricSelector': 'builtin:service.response.time',
        'resolution': '1m',
        'entitySelector': f'entityId({service_id})',
        'from': 'now-30m',
        'to': 'now'
    }

    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()
    data = response.json()

    # Hardcoded for now
    compliance_threshold = 0.99

    # Extract response times (values are in microseconds)
    try:
        series = data['result'][0]['data'][0]
        values = [v for v in series['values'] if v is not None]

        if not values:
            print(f"No data available for service {service_id} in last 5 minutes.")
            return
        
        # Convert from microseconds to milliseconds
        values_ms = [v / 1000 for v in values]

        # Caldular compliance rate
        compliant_count = sum(1 for v in values_ms if v <= threshold)
        compliance_rate = compliant_count / len(values_ms)

        # Calculate average in milliseconds (values are in microseconds)
        avg_response_time = sum(values_ms) / len(values_ms)

        print(f"Service {service_name}: Avg = {avg_response_time:.2f} ms, Compliance = {compliance_rate*100:.2f}%")

        if avg_response_time <= threshold and compliance_rate >= compliance_threshold:
            print(f"Service operating well.")
        else:
            print(f"THRESHOLD EXCEEDED!")

    except (KeyError, IndexError):
        print(f"Unexpected data format received for service {service_id}.")



def output_to_screen_and_file(service_name, service_id, metrics):
    """
    Output response time, total requests, successes, and failures per timestamp to console and a .txt file.
    """
    # Fetch all metrics
    response_time_data = metrics["response_time"]
    request_count_data = metrics["request_count"]
    success_count_data = metrics["success_count"]
    failure_count_data = metrics["failure_count"]

    # Parse metrics into dictionaries
    rt_dict = parse_metric_data(response_time_data)
    req_dict = parse_metric_data(request_count_data)
    success_dict = parse_metric_data(success_count_data)
    failure_dict = parse_metric_data(failure_count_data)

    # Combine all timestamps
    all_timestamps = sorted(set(rt_dict) | set(req_dict) | set(success_dict) | set(failure_dict))

    header = f"Entregando información de consultas sobre el servicio {service_name}"
    print(header)

    # Initialize sums and counter
    success_rate_sum = 0
    failure_rate_sum = 0
    valid_count = 0

    with open(f'{OUTPUT_FOLDER}response_times_{service_name}.txt', 'w', encoding='utf-8') as f:
        f.write(header + "\n\n")

        for ts in all_timestamps:
            time_str = datetime.datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M:%S')
            rt = rt_dict.get(ts)
            req = req_dict.get(ts)
            success = success_dict.get(ts)
            failure = failure_dict.get(ts)

            # Calculate reponse times
            rt_ms = rt / 1000 if rt is not None else None

            # Calculate rates if possible
            success_rate = (success / req * 100) if req and success is not None else None
            failure_rate = (failure / req * 100) if req and failure is not None else None

            # Make sure that both success rate and failure rate have valid values
            # (either both are valid or both aren't). If so, calculate the aggregation
            if success_rate is not None and failure_rate is not None:
                success_rate_sum += success_rate
                failure_rate_sum += failure_rate
                valid_count += 1

            # Build output line
            line = f"{time_str}: Response time: {rt_ms:.2f} ms" if rt_ms is not None else f"{time_str}: Response time: None"
            line += f", Requests: {int(req) if req is not None else 'N/A'}"
            line += f", Successes: {int(success) if success is not None else 'N/A'}"
            line += f", Failures: {int(failure) if failure is not None else 'N/A'}"
            line += f", Success Rate: {success_rate:.2f}%" if success_rate is not None else ", Success Rate: N/A"
            line += f", Failure Rate: {failure_rate:.2f}%" if failure_rate is not None else ", Failure Rate: N/A"

            print(line)
            f.write(line + "\n")
        
        # Calculate and output averages
        f.write("\n")
        print("\n")
        if valid_count > 0:
            avg_success_rate = success_rate_sum / valid_count
            avg_failure_rate = failure_rate_sum / valid_count
            avg_line = (f"Promedio global para {service_name} — "
                        f"Success Rate: {avg_success_rate:.2f}%, "
                        f"Failure Rate: {avg_failure_rate:.2f}%\n")
        else:
            avg_line = f"Promedio global para {service_name} — Sin datos válidos para calcular tasas\n"      
        
        print(avg_line)
        f.write(avg_line + "\n")
        f.write("\nFin de los datos\n")


def output_to_excel(service_name, metrics):
    """
    Create an Excel report with raw data, statistics, daily aggregates, and pivot analysis.
    """

    # Get all the required data from the metrics already obtained for the service
    response_time_data = metrics["response_time"]
    request_count_data = metrics["request_count"]
    success_count_data = metrics["success_count"]
    failure_count_data = metrics["failure_count"]
    
    rt_dict = parse_metric_data(response_time_data)
    req_dict = parse_metric_data(request_count_data)
    success_dict = parse_metric_data(success_count_data)
    failure_dict = parse_metric_data(failure_count_data)

    # series = response_time_data['result'][0]['data'][0]
    # timestamps = series['timestamps']
    # values = series['values']

    # rows = []
    # for ts, val in zip(timestamps, values):
    #     if val is not None:
    #         time_str = datetime.datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M:%S')
    #         ms = val / 1000
    #         threshold = 1 if ms < 3000 else 0
    #         rows.append({
    #             "Timestamp": time_str,
    #             "Response time": round(ms, 2),
    #             "Threshold": threshold
    #         })

    # df = pd.DataFrame(rows)

    # # Create dynamic filename
    timestamp_suffix = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{OUTPUT_FOLDER}{service_name}_response_times_{timestamp_suffix}.xlsx"

    all_timestamps = sorted(set(rt_dict) | set(req_dict) | set(success_dict) | set(failure_dict))

    rows = []
    for ts in all_timestamps:
        time_str = datetime.datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M:%S')
        rt = rt_dict.get(ts)
        req = req_dict.get(ts)
        success = success_dict.get(ts)
        failure = failure_dict.get(ts)

        rt_ms = rt / 1000 if rt is not None else None
        threshold_flag = 1 if rt_ms is not None and rt_ms < 3000 else 0

        success_rate = (success / req * 100) if req and success is not None else None
        failure_rate = (failure / req * 100) if req and failure is not None else None

        rows.append({
            "Timestamp": time_str,
            "Response time": round(rt_ms, 2) if rt_ms is not None else None,
            "Threshold": threshold_flag,
            "Total Requests": int(req) if req is not None else None,
            "Successes": int(success) if success is not None else None,
            "Failures": int(failure) if failure is not None else None,
            "Success Rate (%)": round(success_rate, 2) if success_rate is not None else None,
            "Failure Rate (%)": round(failure_rate, 2) if failure_rate is not None else None
        })

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        raw_numbers_sheet(writer, df, service_name)
        stats_sheet(writer, df)
        daily_averages_sheet(writer, df)
        pivot_table(writer, df)


def raw_numbers_sheet(writer, df, service_name):
    df.to_excel(writer, index=False, sheet_name=service_name)

    # Take care of formatting
    worksheet = writer.sheets[service_name]

    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20  # Requests
    worksheet.column_dimensions['E'].width = 20  # Successes
    worksheet.column_dimensions['F'].width = 20  # Failures
    worksheet.column_dimensions['G'].width = 25  # Success Rate
    worksheet.column_dimensions['H'].width = 25  # Failure Rate



def stats_sheet(writer, df):
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

    # Summary statistics
    avg_response = df["Response time"].mean()
    min_response = df["Response time"].min()
    max_response = df["Response time"].max()
    median_response = df["Response time"].median()
    total_count = len(df)
    above_threshold_pct = (df["Threshold"] == 0).sum() / total_count * 100
    below_threshold_pct = (df["Threshold"] == 1).sum() / total_count * 100

    stats_df = pd.DataFrame({
        "Metric": [
            "Average response time",
            "Fastest response time",
            "Slowest response time",
            "Median response time",
            "Percent above threshold (>= 3000 ms)",
            "Percent below threshold (< 3000 ms)"
        ],
        "Value": [
            round(avg_response, 2),
            round(min_response, 2),
            round(max_response, 2),
            round(median_response, 2),
            above_threshold_pct / 100,
            below_threshold_pct / 100
        ]
    })

    stats_df.to_excel(writer, index=False, sheet_name="Stats")
    worksheet = writer.sheets["Stats"]

    # Column widths
    worksheet.column_dimensions['A'].width = 50
    worksheet.column_dimensions['B'].width = 25

    # Styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=16)
    center_align = Alignment(horizontal='center', vertical='center')

    for row in worksheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            elif cell.column_letter == 'B':
                cell.alignment = center_align

    for row in worksheet.iter_rows(min_row=2, max_col=2, max_row=7):
        metric_cell, value_cell = row
        if "Percent" in metric_cell.value:
            value_cell.number_format = '0.00%'
        else:
            value_cell.number_format = '#,##0.0'


def daily_averages_sheet(writer, df):
    df["Date"] = pd.to_datetime(df["Timestamp"]).dt.date
    daily_stats_df = df.groupby("Date").agg({
        "Response time": ["mean", "min", "max", "median", 
                           lambda x: (x >= 3000).mean() * 100, 
                           lambda x: (x < 3000).mean() * 100]
    })
    daily_stats_df.columns = [
        "Average response time",
        "Fastest response time",
        "Slowest response time",
        "Median response time",
        "Percent above threshold (>= 3000 ms)",
        "Percent below threshold (< 3000 ms)"
    ]
    daily_stats_df = daily_stats_df.reset_index().round(2)
    daily_stats_df.to_excel(writer, index=False, sheet_name="Daily averages")

    # Take care of formatting
    worksheet = writer.sheets["Daily averages"]

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 25
    worksheet.column_dimensions['E'].width = 25
    worksheet.column_dimensions['F'].width = 30
    worksheet.column_dimensions['G'].width = 30

    center_align = Alignment(horizontal='center', vertical='center')

    for row in worksheet:
        for cell in row:
            cell.alignment = center_align


def pivot_table(writer, df):
    df["Date"] = pd.to_datetime(df["Timestamp"]).dt.date
    df["Hour"] = pd.to_datetime(df["Timestamp"]).dt.hour
    pivot_df = df.groupby(["Date", "Hour"]).agg(
        total=("Threshold", "count"),
        under_threshold=("Threshold", "sum")
    ).reset_index()

    pivot_df["over_threshold"] = pivot_df["total"] - pivot_df["under_threshold"]
    pivot_df["% under"] = (pivot_df["under_threshold"] / pivot_df["total"]) * 100
    pivot_df["% over"] = 100 - pivot_df["% under"]

    pivot_df.to_excel(writer, sheet_name="Pivot", index=False)

    # Take care of formatting
    worksheet = writer.sheets["Pivot"]

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 25
    worksheet.column_dimensions['E'].width = 25
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20

    center_align = Alignment(horizontal='center', vertical='center')

    for row in worksheet:
        for cell in row:
            cell.alignment = center_align



# --- Entry Point ---

def main():

    THRESHOLD_MS = 3000  # example threshold in milliseconds, adjust as needed
    print("\n")

    for service_name, service_id in SERVICE_MAP.items():
        try:
            
            #data = get_service_performance(service_id, 'builtin:service.response.time')
            metrics = get_all_service_metrics(service_id)
            print (f"Obtained data for service {service_name}. Will create Excel file...\n")
            # output_to_screen_and_file(service_name, service_id, metrics)
            output_to_excel(service_name, metrics)
            print (f"Excel file for service {service_name} created successfully\n")
            # poll_response_time(service_name, service_id, THRESHOLD_MS)
        except Exception as e:
            print(f"Error processing service {service_name}: {e}")
    
    print("\n")

    # THRESHOLD_MS = 3000  # example threshold in milliseconds, adjust as needed
    # print("\nPolling started. Press Ctrl+C to stop.\n")

    # try:
    #     while True:
    #         for service_name, service_id in SERVICE_MAP.items():
    #             try:
    #                 poll_response_time(service_name, service_id, THRESHOLD_MS)
    #             except Exception as e:
    #                 print(f"Error processing service {service_name}: {e}")
    #         print("-" * 60)
    #         time.sleep(10)  # wait 30 seconds before polling again

    # except KeyboardInterrupt:
    #     print("\nPolling stopped by user.") 


if __name__ == "__main__":
    main()
