import requests
import datetime
import pandas as pd
from dotenv import load_dotenv
import os
import time
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Configuration ---


load_dotenv()
API_TOKEN = os.getenv("DYNATRACE_API_TOKEN")

ENV_ID = 'kae68552'
OUTPUT_FOLDER = "output/"

# Service name → ID mapping
SERVICE_MAP = {
    "ComercioTransaccionesController": "SERVICE-FD9343224D905203",
    "AbonosController": "SERVICE-123E236BA4855F4A"
    # "TransaccionesMultiserviciosController": "SERVICE-E77A12909369A78D"
}

DATABASE_MAP = {
    "mc_tlog_historia": "SERVICE-0604902B4FAD42B6"
}

DB_METRICS = {
    "Total Connections": "builtin:service.dbconnections.total",
    "Failure Rate": "builtin:service.dbconnections.failureRate",
    "Success Rate": "builtin:service.dbconnections.successRate",
    "Failed Connections": "builtin:service.dbconnections.failure"
}


# --- Functions ---


# Methods to get service performance
def get_service_performance(service_id, metric_selector):
    """
    Fetch metrics for a given service from Dynatrace.
    """
    url = f'https://{ENV_ID}.live.dynatrace.com/api/v2/metrics/query'
    headers = {
        'Authorization': f'Api-Token {API_TOKEN}'
    }
    params = {
        'metricSelector': metric_selector,
        'resolution': '1h',
        'entitySelector': f'entityId({service_id})',
        'from': 'now-365d',
        'to': 'now'
    }
    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()
    return response.json()

def get_all_service_metrics(service_id):
    return {
        "response_time": get_service_performance(service_id, 'builtin:service.response.time'),
        "request_count": get_service_performance(service_id, 'builtin:service.requestCount.total'),
        "success_count": get_service_performance(service_id, 'builtin:service.errors.server.successCount'),
        "failure_count": get_service_performance(service_id, 'builtin:service.errors.server.count'),
    }

def get_metric_data(service_id, metric_key):
    url = f"https://{ENV_ID}.live.dynatrace.com/api/v2/metrics/query"
    headers = {
        "Authorization": f"Api-Token {API_TOKEN}"
    }
    params = {
        "metricSelector": metric_key,
        "entitySelector": f"entityId({service_id})",
        "from": "now-7d",  # Or adjust timeframe
        "to": "now",
        "resolution": "1m"
    }
    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()
    return response.json()

def parse_metric_timeseries(data):
    series = data['result'][0]['data'][0]
    timestamps = series['timestamps']
    values = series['values']
    return {ts: val for ts, val in zip(timestamps, values)}

def output_database_metrics_to_excel(db_name, service_id):
    print(f"Gathering DB metrics for {db_name}...")
    metrics_data = {}

    # Fetch and parse each metric
    for label, metric_key in DB_METRICS.items():
        try:
            raw_data = get_metric_data(service_id, metric_key)
            metrics_data[label] = parse_metric_timeseries(raw_data)
        except Exception as e:
            print(f"Error fetching {label} for {db_name}: {e}")
            metrics_data[label] = {}

    # Merge all timestamps
    all_timestamps = sorted(set(ts for d in metrics_data.values() for ts in d))

    rows = []
    for ts in all_timestamps:
        time_str = datetime.datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M:%S')
        row = {
            "Timestamp": time_str
        }
        for label in DB_METRICS.keys():
            val = metrics_data[label].get(ts)
            row[label] = round(val, 4) if val is not None else (100 if label == "Success Rate" else 0)
        rows.append(row)

    df = pd.DataFrame(rows)

    # Add new metric for "lagged" failes connections
    df["Lagged Failed Connections"] = df["Failed Connections"].shift(1).rolling(window=10).sum()

    # Write to Excel
    timestamp_suffix = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{OUTPUT_FOLDER}{db_name}_db_metrics_{timestamp_suffix}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Raw DB Metrics", index=False)

        # Format columns
        worksheet = writer.sheets["Raw DB Metrics"]
        for col in worksheet.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Compute correlations (ignore NaNs)
    try:
        corr_success = df["Total Connections"].corr(df["Success Rate"])
        corr_failure = df["Total Connections"].corr(df["Failure Rate"])
        corr_failed_conn = df["Failed Connections"].corr(df["Failure Rate"])
        corr_lag_failed_conn = df["Lagged Failed Connections"].corr(df["Failure Rate"])

        print(f"Correlation (Total Connections ↔ Success Rate): {corr_success:.4f}")
        print(f"Correlation (Total Connections ↔ Failure Rate): {corr_failure:.4f}")
        print(f"Correlation (Failed Connections ↔ Failure Rate): {corr_failed_conn}")
        print(f"Correlation (Lagged Failed Connections ↔ Failure Rate): {corr_lag_failed_conn}")
    except Exception as e:
        print(f"Could not compute correlations: {e}")

def parse_metric_data(data):
    series = data['result'][0]['data'][0]
    timestamps = series['timestamps']
    values = series['values']
    return {ts: val for ts, val in zip(timestamps, values)}

def list_request_names_for_service(service_name, service_id):
    print(f"Service name: {service_name}\n")
    url = f"https://{ENV_ID}.live.dynatrace.com/api/v2/metrics/query"
    headers = {
        "Authorization": f"Api-Token {API_TOKEN}"
    }

    params = {
        "metricSelector": 'builtin:service.response.time',
        "entitySelector": f"entityId({service_id})",
        "from": "now-15m",
        "to": "now",
        "resolution": "1m"
    }

    try:
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        data = response.json()

        print(f"\n=== Request Names for Service '{service_name}' ===")
        # for series in data.get("result", [])[0].get("data", []):
        #     dimension = series.get("dimensions", [])[0]
        #     if dimension:
        #         print(f"- {dimension}")
        print(data)
        print("\n")

    except Exception as e:
        print(f"Error fetching request names for {service_name}: {e}")

def get_metric_dimensions_for_services(services: dict):
    url = f"https://{ENV_ID}.live.dynatrace.com/api/v2/metrics/{'builtin:service.response.time'}"
    headers = {
        "Authorization": f"Api-Token {API_TOKEN}"
    }

    for service_name, service_id in services.items():
        print(f"\n=== Service: {service_name} (ID: {service_id}) ===")

        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            data = response.json()

            dimensions = data.get("dimensionDefinitions", [])
            if not dimensions:
                print("No dimensions found.")
                continue

            print(f"Dimensions for metric '{'builtin:service.response.time'}':")
            for d in dimensions:
                key = d.get("key")
                name = d.get("name")
                print(f"- {key} ({name})")

        except Exception as e:
            print(f"Error fetching dimensions for {service_name}: {e}")


def poll_response_time(service_name, service_id, threshold):
    """
    Fetch average response time over the last 5 minutes and compare against threshold.
    """

    # Get current timestamp for logging
    from datetime import datetime
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

     # Initialize log file if first run
    try:
        with open(f'{OUTPUT_FOLDER}polling_results.txt', 'x') as f:
            f.write(f"Polling started at: {current_time}\n")
    except FileExistsError:
        pass  # File already exists, no need to initialize

    # Query last 5 minutes data from Dynatrace
    url = f'https://{ENV_ID}.live.dynatrace.com/api/v2/metrics/query'
    headers = {
        'Authorization': f'Api-Token {API_TOKEN}'
    }
    params = {
        'metricSelector': 'builtin:service.response.time',
        'resolution': '1m',
        'entitySelector': f'entityId({service_id})',
        'from': 'now-30m',
        'to': 'now'
    }

    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()
    data = response.json()

    # Hardcoded for now
    compliance_threshold = 0.99

    # Extract response times (values are in microseconds)
    try:
        series = data['result'][0]['data'][0]
        values = [v for v in series['values'] if v is not None]

        if not values:
            msg = f"{current_time} - No data available for service {service_id} in last 5 minutes."
            print(msg)
            with open('polling_results.txt', 'a') as f:
                f.write(msg + '\n')
            return
        
        # Convert from microseconds to milliseconds
        values_ms = [v / 1000 for v in values]

        # Calculate compliance rate
        compliant_count = sum(1 for v in values_ms if v <= threshold)
        compliance_rate = compliant_count / len(values_ms)

        # Calculate average in milliseconds
        avg_response_time = sum(values_ms) / len(values_ms)

        # Prepare log message
        status = "OK" if (avg_response_time <= threshold and compliance_rate >= compliance_threshold) else "THRESHOLD EXCEEDED"
        msg = (f"{current_time} - Service {service_name}: "
               f"Avg = {avg_response_time:.2f} ms, "
               f"Compliance = {compliance_rate*100:.2f}%, "
               f"Status = {status}")

        print(msg)
        
        # Append to log file
        with open(f'{OUTPUT_FOLDER}polling_results.txt', 'a') as f:
            f.write(msg + '\n')

    except (KeyError, IndexError) as e:
        msg = f"{current_time} - Error processing data for service {service_id}: {str(e)}"
        print(msg)
        with open('polling_results.txt', 'a') as f:
            f.write(msg + '\n')



def output_to_screen_and_file(service_name, service_id, metrics):
    """
    Output response time, total requests, successes, and failures per timestamp to console and a .txt file.
    """
    # Fetch all metrics
    response_time_data = metrics["response_time"]
    request_count_data = metrics["request_count"]
    success_count_data = metrics["success_count"]
    failure_count_data = metrics["failure_count"]

    # Parse metrics into dictionaries
    rt_dict = parse_metric_data(response_time_data)
    req_dict = parse_metric_data(request_count_data)
    success_dict = parse_metric_data(success_count_data)
    failure_dict = parse_metric_data(failure_count_data)

    # Combine all timestamps
    all_timestamps = sorted(set(rt_dict) | set(req_dict) | set(success_dict) | set(failure_dict))

    header = f"Entregando información de consultas sobre el servicio {service_name}"
    print(header)

    # Initialize sums and counter
    success_rate_sum = 0
    failure_rate_sum = 0
    valid_count = 0

    with open(f'{OUTPUT_FOLDER}response_times_{service_name}.txt', 'w', encoding='utf-8') as f:
        f.write(header + "\n\n")

        for ts in all_timestamps:
            time_str = datetime.datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M:%S')
            rt = rt_dict.get(ts)
            req = req_dict.get(ts)
            success = success_dict.get(ts)
            failure = failure_dict.get(ts)

            # Calculate reponse times
            rt_ms = rt / 1000 if rt is not None else None

            # Calculate rates if possible
            success_rate = (success / req * 100) if req and success is not None else None
            failure_rate = (failure / req * 100) if req and failure is not None else None

            # Make sure that both success rate and failure rate have valid values
            # (either both are valid or both aren't). If so, calculate the aggregation
            if success_rate is not None and failure_rate is not None:
                success_rate_sum += success_rate
                failure_rate_sum += failure_rate
                valid_count += 1

            # Build output line
            line = f"{time_str}: Response time: {rt_ms:.2f} ms" if rt_ms is not None else f"{time_str}: Response time: None"
            line += f", Requests: {int(req) if req is not None else 'N/A'}"
            line += f", Successes: {int(success) if success is not None else 'N/A'}"
            line += f", Failures: {int(failure) if failure is not None else 'N/A'}"
            line += f", Success Rate: {success_rate:.2f}%" if success_rate is not None else ", Success Rate: N/A"
            line += f", Failure Rate: {failure_rate:.2f}%" if failure_rate is not None else ", Failure Rate: N/A"

            print(line)
            f.write(line + "\n")
        
        # Calculate and output averages
        f.write("\n")
        print("\n")
        if valid_count > 0:
            avg_success_rate = success_rate_sum / valid_count
            avg_failure_rate = failure_rate_sum / valid_count
            avg_line = (f"Promedio global para {service_name} — "
                        f"Success Rate: {avg_success_rate:.2f}%, "
                        f"Failure Rate: {avg_failure_rate:.2f}%\n")
        else:
            avg_line = f"Promedio global para {service_name} — Sin datos válidos para calcular tasas\n"      
        
        print(avg_line)
        f.write(avg_line + "\n")
        f.write("\nFin de los datos\n")


def output_to_excel(service_name, metrics):
    """
    Create an Excel report with raw data, statistics, daily aggregates, and pivot analysis.
    """

    # Get all the required data from the metrics already obtained for the service
    response_time_data = metrics["response_time"]
    request_count_data = metrics["request_count"]
    success_count_data = metrics["success_count"]
    failure_count_data = metrics["failure_count"]
    
    rt_dict = parse_metric_data(response_time_data)
    req_dict = parse_metric_data(request_count_data)
    success_dict = parse_metric_data(success_count_data)
    failure_dict = parse_metric_data(failure_count_data)

    # # Create dynamic filename
    timestamp_suffix = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{OUTPUT_FOLDER}{service_name}_response_times_{timestamp_suffix}.xlsx"

    all_timestamps = sorted(set(rt_dict) | set(req_dict) | set(success_dict) | set(failure_dict))

    rows = []
    for ts in all_timestamps:
        time_str = datetime.datetime.fromtimestamp(ts / 1000).strftime('%Y-%m-%d %H:%M:%S')
        rt = rt_dict.get(ts)
        req = req_dict.get(ts)
        success = success_dict.get(ts)
        failure = failure_dict.get(ts)

        rt_ms = rt / 1000 if rt is not None else None
        threshold_flag = 1 if rt_ms is not None and rt_ms < 3000 else 0

        success_rate = (success / req) if req and success is not None else None
        failure_rate = (failure / req) if req and failure is not None else None

        rows.append({
            "Timestamp": time_str,
            "Response time": round(rt_ms, 2) if rt_ms is not None else None,
            "Threshold": threshold_flag if rt_ms is not None else 1,
            "Total Requests": int(req) if req is not None else 0,
            "Successes": int(success) if success is not None else 0,
            "Failures": int(failure) if failure is not None else 0,
            "Success Rate (%)": round(success_rate, 2) if success_rate is not None else 1,
            "Failure Rate (%)": round(failure_rate, 2) if failure_rate is not None else 0
        })

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        raw_numbers_sheet(writer, df, service_name)
        stats_sheet(writer, df)
        daily_averages_sheet(writer, df)
        pivot_table(writer, df)


def raw_numbers_sheet(writer, df, service_name):
    # Write data to Excel first
    df.to_excel(writer, index=False, sheet_name=service_name)
    
    # Apply formatting
    worksheet = writer.sheets[service_name]
    
    # Define styles (consistent with other sheets)
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), 
                  right=Side(style='thin'), 
                  top=Side(style='thin'), 
                  bottom=Side(style='thin'))
    
    # Set column widths (as you specified)
    column_widths = {
        'A': 30,  # Timestamp
        'B': 20,  # Response time
        'C': 20,  # Threshold
        'D': 20,  # Requests
        'E': 20,  # Successes
        'F': 20,  # Failures
        'G': 25,  # Success Rate
        'H': 25   # Failure Rate
    }
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # Apply header style
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Apply data styles
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = center_alignment
            
            # Apply number formatting based on column
            col_letter = cell.column_letter
            if col_letter == 'A':  # Timestamp
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
            elif col_letter == 'B':  # Response time
                cell.number_format = '0.00'
            elif col_letter == 'C':  # Threshold
                cell.number_format = '0'
            elif col_letter in ('D', 'E', 'F'):  # Requests/Successes/Failures
                cell.number_format = '0'
            elif col_letter in ('G', 'H'):  # Rates
                cell.number_format = '0.00%'
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'


def stats_sheet(writer, df):
    # Calculate statistics with full precision
    avg_response = float(df["Response time"].mean())
    min_response = float(df["Response time"].min())
    max_response = float(df["Response time"].max())
    median_response = float(df["Response time"].median())
    total_count = len(df)
    
    # Calculate threshold percentages
    above_threshold_count = (df["Threshold"] == 0).sum()
    below_threshold_count = (df["Threshold"] == 1).sum()
    above_threshold_pct = above_threshold_count / total_count  # 0-1 scale
    below_threshold_pct = below_threshold_count / total_count  # 0-1 scale
    
    total_requests = int(df["Total Requests"].sum())
    total_successes = int(df["Successes"].sum())
    total_failures = int(df["Failures"].sum())
    
    # Calculate rates
    overall_success_rate = float(total_successes / total_requests) if total_requests else 0.0
    overall_failure_rate = float(total_failures / total_requests) if total_requests else 0.0

    # Create DataFrame with metrics
    stats_df = pd.DataFrame({
        "Metric": [
            "Average response time (ms)",
            "Fastest response time (ms)",
            "Slowest response time (ms)",
            "Median response time (ms)",
            "Percent above threshold (>= 3000 ms)",
            "Percent below threshold (< 3000 ms)",
            "Total number of requests",
            "Total successful requests",
            "Total requests with error",
            "Overall success rate",
            "Overall failure rate"
        ],
        "Value": [
            avg_response,
            min_response,
            max_response,
            median_response,
            above_threshold_pct,  # Will show as percentage
            below_threshold_pct,   # Will show as percentage
            total_requests,
            total_successes,
            total_failures,
            overall_success_rate,  # Will show as percentage
            overall_failure_rate   # Will show as percentage
        ]
    })

    # Write to Excel
    stats_df.to_excel(writer, index=False, sheet_name="Stats")
    worksheet = writer.sheets["Stats"]

    # Set column widths
    worksheet.column_dimensions['A'].width = 35  # Metric column
    worksheet.column_dimensions['B'].width = 25  # Value column

    # Define styles
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), 
                  right=Side(style='thin'), 
                  top=Side(style='thin'), 
                  bottom=Side(style='thin'))

    # Apply header style
    for cell in worksheet[1]:  # Header row
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Apply formatting by absolute row position
    for row_idx in range(2, worksheet.max_row + 1):
        metric_cell = worksheet.cell(row=row_idx, column=1)
        value_cell = worksheet.cell(row=row_idx, column=2)
        
        # Apply borders and alignment to all cells
        metric_cell.border = border
        value_cell.border = border
        value_cell.alignment = center_align
        
        # Force formatting based on Excel row number (1-based)
        if row_idx in [6, 7]:  # Percent threshold rows (rows 6-7 in Excel)
            value_cell.number_format = '0.00%'
        elif row_idx in [11, 12]:  # Rate rows (rows 11-12 in Excel)
            value_cell.number_format = '0.00%'
        elif row_idx in [2, 3, 4, 5]:  # Response time rows (rows 2-5)
            value_cell.number_format = '0.00'
        else:  # Count rows (rows 8-10)
            value_cell.number_format = '#,##0'


def daily_averages_sheet(writer, df):
    # Convert timestamp to date and create Date column
    df["Date"] = pd.to_datetime(df["Timestamp"]).dt.date
    
    # Perform aggregations
    daily_stats_df = df.groupby("Date").agg({
        "Response time": ["mean", "min", "max", "median"],
        "Threshold": "mean",  # This gives the proportion below threshold
        "Total Requests": "sum",
        "Successes": "sum",
        "Failures": "sum"
    })

    # Calculate percentages
    percent_below = daily_stats_df[("Threshold", "mean")]
    percent_above = 1 - percent_below

    # Compute success/failure rates
    daily_stats_df[("Success Rate", "")] = (
        daily_stats_df[("Successes", "sum")] / daily_stats_df[("Total Requests", "sum")]
    )
    daily_stats_df[("Failure Rate", "")] = (
        daily_stats_df[("Failures", "sum")] / daily_stats_df[("Total Requests", "sum")]
    )

    # Flatten the multi-index columns
    daily_stats_df.columns = [
        "Average response time",
        "Fastest response time",
        "Slowest response time",
        "Median response time",
        "Percent below threshold (< 3000 ms)",
        "Total requests",
        "Total successes",
        "Total errors",
        "Overall success rate (%)",
        "Overall failure rate (%)"
    ]
    
    # Add the above-threshold percentage
    daily_stats_df["Percent above threshold (>= 3000 ms)"] = percent_above

    # Reset index to make Date a column and define column order
    daily_stats_df = daily_stats_df.reset_index()
    column_order = [
        "Date",
        "Average response time",
        "Fastest response time",
        "Slowest response time",
        "Median response time",
        "Percent below threshold (< 3000 ms)",
        "Percent above threshold (>= 3000 ms)",
        "Total requests",
        "Total successes",
        "Total errors",
        "Overall success rate (%)",
        "Overall failure rate (%)"
    ]
    
    # Select and reorder columns
    daily_stats_df = daily_stats_df[column_order]
    
    # Write to Excel
    daily_stats_df.to_excel(writer, index=False, sheet_name="Daily averages")

    # Take care of formatting
    worksheet = writer.sheets["Daily averages"]

    # Column-specific widths
    column_widths = [
        18, 20, 20, 20, 20, 20, 18, 20, 20, 18, 18, 18
    ]

    # Set column widths
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Define styles
    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

    # Number formats
    time_format = '0.00'
    count_format = '0'
    percent_format = '0.00%'

    # Apply header style
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # Apply data styles
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            
            # Apply specific number formats based on column
            col_letter = cell.column_letter
            col_idx = cell.column
            
            # Date column (A)
            if col_idx == 1:
                cell.number_format = 'YYYY-MM-DD'
            
            # Time measurement columns (B-E)
            elif 2 <= col_idx <= 5:
                cell.number_format = time_format
            
            # Percentage columns (F,G,J,K)
            elif col_idx in [6, 7, 11, 12]:
                cell.number_format = percent_format
            
            # Count columns (H,I)
            elif col_idx in [8, 9, 10]:
                cell.number_format = count_format



def pivot_table(writer, df):
    df["Date"] = pd.to_datetime(df["Timestamp"]).dt.date
    df["Hour"] = pd.to_datetime(df["Timestamp"]).dt.hour
    
    # Enhanced aggregation
    pivot_df = df.groupby(["Date", "Hour"]).agg(
        total_samples=("Threshold", "count"),
        under_threshold=("Threshold", "sum"),
        total_requests=("Total Requests", "sum"),
        total_successes=("Successes", "sum"),
        total_failures=("Failures", "sum")
    ).reset_index()

    # Convert hour numbers to time intervals
    pivot_df["Time Interval"] = pivot_df["Hour"].apply(
        lambda x: f"{x}:00 - {x+1}:00" if x < 23 else "23:00 - 0:00"
    )
    
    # Calculate derived metrics
    pivot_df["over_threshold"] = pivot_df["total_samples"] - pivot_df["under_threshold"]
    pivot_df["% under"] = pivot_df["under_threshold"] / pivot_df["total_samples"]
    pivot_df["% over"] = 1 - pivot_df["% under"]
    pivot_df["success_rate"] = pivot_df["total_successes"] / pivot_df["total_requests"]
    pivot_df["failure_rate"] = pivot_df["total_failures"] / pivot_df["total_requests"]

    # Reorder and select columns
    pivot_df = pivot_df[[
        "Date", "Time Interval",
        "total_samples",
        "under_threshold", "over_threshold", "% under", "% over",
        "total_requests",
        "total_successes", "total_failures",
        "success_rate", "failure_rate"
    ]]
    
    # Write to Excel
    pivot_df.to_excel(writer, sheet_name="Pivot", index=False)
    
    # Apply formatting
    if isinstance(writer, pd.ExcelWriter):
        worksheet = writer.sheets["Pivot"]
        
        # Define styles
        header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))
        
        # Set column widths (updated for time interval column)
        column_widths = [12, 18, 12, 15, 15, 12, 12, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Apply header style
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Apply data styles
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = border
                
                # Apply number formatting
                col_idx = cell.column
                if col_idx in [6, 7, 11, 12]:  # Percentage columns (shifted right by 1)
                    cell.number_format = '0.00%'
                elif col_idx in [3, 4, 5, 8, 9, 10]:  # Count columns
                    cell.number_format = '0'
                elif col_idx == 1:  # Date column
                    cell.number_format = 'YYYY-MM-DD'
                # Time Interval column (col_idx=2) keeps text format


# --- Entry Point ---

def main():

    THRESHOLD_MS = 3000  # example threshold in milliseconds, adjust as needed
    print("\n")

    for service_name, service_id in SERVICE_MAP.items():
        try:
            
            #data = get_service_performance(service_id, 'builtin:service.response.time')
            metrics = get_all_service_metrics(service_id)
            print (f"Obtained data for service {service_name}. Will create Excel file...\n")
            output_to_screen_and_file(service_name, service_id, metrics)
            output_to_excel(service_name, metrics)
            print (f"Excel file for service {service_name} created successfully\n")
        except Exception as e:
            print(f"Error processing service {service_name}: {e}")
    
    print("\n")

    # print("Ahora analizar las DB...")
    # for db_name, service_id in DATABASE_MAP.items():
    #     output_database_metrics_to_excel(db_name, service_id)

    # THRESHOLD_MS = 3000  # example threshold in milliseconds, adjust as needed
    print("\nPolling started. Press Ctrl+C to stop.\n")

    # try:
    #     while True:
    #         for service_name, service_id in SERVICE_MAP.items():
    #             try:
    #                 poll_response_time(service_name, service_id, THRESHOLD_MS)
    #             except Exception as e:
    #                 print(f"Error processing service {service_name}: {e}")
    #         print("-" * 60)
    #         time.sleep(30)  # wait 30 seconds before polling again

    # except KeyboardInterrupt:
    #     print("\nPolling stopped by user.") 
    # for service_name, service_id in SERVICE_MAP.items():
    #     list_request_names_for_service(service_name, service_id)



if __name__ == "__main__":
    main()
