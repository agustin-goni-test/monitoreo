import openpyxl.styles
import openpyxl.utils
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.text import Font, CharacterProperties
from openpyxl.workbook import Workbook
from openpyxl.chart.label import DataLabel
import openpyxl
import subprocess
from typing import List, Tuple

class ExcelEnhancer:
    def __init__(self, file_name):
        self.original_file_name = file_name
        if not file_name.endswith('.xlsx'):
            raise ValueError("File name must end with '.xlsx'.")
        self.enhanced_file_name = None

    def get_output_file_name(self):
        """Generate and set the output file name for the enhanced Excel file inside an 'output' folder."""
        base_name = os.path.basename(self.original_file_name)
        name, ext = os.path.splitext(base_name)
        output_dir = os.path.join(os.path.dirname(self.original_file_name), "enhanced")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        self.enhanced_file_name = os.path.join(output_dir, f"{name}_ENH{ext}")


    def set_file_name(self, file_name):
        """Set the file name for the Excel file."""
        if not isinstance(file_name, str):
            raise ValueError("File name must be a string.")
        if not file_name.endswith('.xlsx'):
            raise ValueError("File name must end with '.xlsx'.")
        self.original_file_name = file_name

    def add_context_columns(self, output_file_name=None):
        """Add context columns to the Excel file and save it."""

        # Load the Excel file
        try:   
            df = pd.read_excel(self.original_file_name)
        except FileNotFoundError:
            raise FileNotFoundError(f"The file {self.original_file_name} does not exist.")
        except Exception as e:
            raise Exception(f"An error occurred while reading the file: {str(e)}")
        
        # Check column A is a timestamp column
        timestamp_col = df.columns[0]

        # Convert timestamp column to datetime if not already
        if not pd.api.types.is_datetime64_any_dtype(df[timestamp_col]):
            try:
                df[timestamp_col] = pd.to_datetime(df[timestamp_col], errors='coerce')
            except Exception as e:
                raise ValueError(f"Column {timestamp_col} cannot be converted to datetime: {str(e)}")
            
        # Add context columns
        df.insert(1, "Date", df[timestamp_col].dt.date.astype(str))
        df.insert(2, "Month", df[timestamp_col].dt.month)
        df.insert(3, "Day", df[timestamp_col].dt.day)
        df.insert(4, "Weekday", df[timestamp_col].dt.strftime('%a'))
        df.insert(5, "Hour", df[timestamp_col].dt.hour)

        if self.enhanced_file_name is None:
            self.get_output_file_name()

        df.to_excel(self.enhanced_file_name, index=False)
        print(f"Enhanced file saved as {self.enhanced_file_name}")


    def insert_time_pivot(self):
        """Insert a pivot table based on the time context columns."""
        if self.enhanced_file_name is None:
            raise ValueError("Enhanced file name is not set. Call add_context_columns() first.")

        # Load the enhanced Excel file
        try:
            df = pd.read_excel(self.enhanced_file_name)
        except FileNotFoundError:
            raise FileNotFoundError(f"The file {self.enhanced_file_name} does not exist.")
        except Exception as e:
            raise Exception(f"An error occurred while reading the file: {str(e)}")
        
        # Check for Date column
        if "Date" not in df.columns:
            raise ValueError("The enhanced file does not contain a 'Date' column. Please run add_context_columns() first.")
        
        # # Find the first column that contains "Time" in its name
        # time_column = next((col for col in df.columns if "Time" in col and "Timestamp" not in col), None)

        # if not time_column:
        #     raise ValueError("No column containing 'Time' found in the enhanced file.")
        
        # # Create a pivot table
        # pivot_df = df.groupby("Date", as_index=False)[time_column].mean()
        # pivot_df.rename(columns={time_column: "Average Time"}, inplace=True)

        # Create result data frame to include all colums
        result_df = df[['Date']].drop_duplicates().reset_index(drop=True)

        # Find time columns
        for col in df.columns:
            if "Time" in col and "Timestamp" not in col and "Compliance" not in col:
                time_column = col
                
                # Rename the column for the target pivot table
                if "Response Time" in time_column:
                    new_column_name = col.split("Response Time")[0] + "Average Time (s)"
                else:
                    new_column_name = "Average Time (s)"

                # Summarize by date using the mean of the time column
                dimension_df = df.groupby("Date", as_index=False)[time_column].mean()
                dimension_df.rename(columns={time_column: new_column_name})

                # Merge with result data frame
                result_df = result_df.merge(
                    dimension_df,
                    on="Date",
                    how="left"
                )   
               

        # Write to a new sheet in the enhanced file
        with pd.ExcelWriter(self.enhanced_file_name, engine='openpyxl', mode='a') as writer:
            result_df.to_excel(writer, sheet_name='Time Pivot', index=False)

        print(f"Pivot table added to {self.enhanced_file_name} in 'Time Pivot' sheet.")


    def insert_compliance_pivot(self):
        if self.enhanced_file_name is None:
            raise ValueError("Enhanced file name is not set. Call add_context_columns() first.")
        
         # Load the enhanced Excel file
        try:
            df = pd.read_excel(self.enhanced_file_name)
        except FileNotFoundError:
            raise FileNotFoundError(f"The file {self.enhanced_file_name} does not exist.")
        except Exception as e:
            raise Exception(f"An error occurred while reading the file: {str(e)}")
        
        # Check for Date column
        if "Date" not in df.columns:
            raise ValueError("The enhanced file does not contain a 'Date' column. Please run add_context_columns() first.")
        
        # Create result data frame to include all colums
        result_df = df[['Date']].drop_duplicates().reset_index(drop=True)

        for col in df.columns:
            if "Compliance" in col or "Rate" in col:
                # Determine which suffix to use
                suffix = "Compliance (%)" if "Compliance" in col else "Rate (%)"

                # Extract the prefix before either "Response Time" or "Time"
                if "Response Time" in col:
                    new_column_name = col.split("Response Time")[0] + suffix
                elif "Time" in col:
                    new_column_name = col.split("Time")[0] + suffix
                else:
                    new_column_name = col

                # Summarize by date using the mean of the compliance column
                dimension_df = df.groupby("Date", as_index=False)[col].mean()
                dimension_df.rename(columns={col: new_column_name}, inplace=True)

                # Merge with result data frame
                result_df = result_df.merge(
                    dimension_df,
                    on="Date",
                    how="left"
                )

        # Write to a new sheet in the enhanced file
        with pd.ExcelWriter(self.enhanced_file_name, engine='openpyxl', mode='a') as writer:
            result_df.to_excel(writer, sheet_name='Comp Pivot', index=False)

        print(f"Pivot table added to {self.enhanced_file_name} in 'Comp Pivot' sheet.")


    def insert_count_pivot(self):
        """Insert a pivot table based on the count context columns."""
        self.insert_count_pivot_by_column("Date")
        self.insert_count_pivot_by_column("Hour")
        self.insert_count_pivot_by_column("Weekday")


    def insert_count_pivot_by_column(self, time_column="Date"):
        if self.enhanced_file_name is None:
            raise ValueError("Enhanced file name is not set. Call add_context_columns() first.")
        
         # Load the enhanced Excel file
        try:
            df = pd.read_excel(self.enhanced_file_name)
        except FileNotFoundError:
            raise FileNotFoundError(f"The file {self.enhanced_file_name} does not exist.")
        except Exception as e:
            raise Exception(f"An error occurred while reading the file: {str(e)}")
        
        pairs = []
        
        result_df, sheet_name = self.calculate_count_pivot_by_column(df, time_column)
        pairs.append((result_df, sheet_name))

        self.write_pivot_sheets(pairs)

        
        # # Chech the required columns exist in the Excel file
        # required_columns = ['Date', time_column]
        # missing = [col for col in required_columns if col not in df.columns]
        # if missing:
        #     raise ValueError(f"The enhanced file is missing required columns: {', '.join(missing)}")
        
        # # Create result data frame to include all colums
        # result_df = pd.DataFrame({time_column: sorted(df[time_column].unique())})

        # for col in df.columns:
        #     if "Count" in col or "Total" in col:

        #         # Maintain the original column name for clarity, no need for changes

        #         # Separate different cases: time_column is "Date", or something else
        #         if time_column == "Date":
        #             dimension_df = df.groupby("Date", as_index=False)[col].sum()
        #         else:
        #             # Two steps to summarize: step 1 - Sum by Date and time_column ("Hour", "Weekday", etc.)
        #             daily_totals_df = df.groupby(["Date", time_column], as_index=False)[col].sum()

        #             # Step 2 - Average across dates for each time_column
        #             dimension_df = daily_totals_df.groupby(time_column, as_index=False)[col].mean()

        #         # Convert to int (as these are counts)
        #         dimension_df[col] = dimension_df[col].astype(int)

        #         # Merge with result data frame
        #         result_df = result_df.merge(
        #             dimension_df,
        #             on=time_column,
        #             how="left"
        #         )
        
        # # Define the name for the new sheet according to situation
        # sheet_name = 'Count Pivot' if time_column == "Date" else f'Count Pivot - {time_column}'
        # if len(sheet_name) > 31:  # Excel sheet names must be <= 31 characters
        #     sheet_name = sheet_name[:31]


        # Write to a new sheet in the enhanced file
        # with pd.ExcelWriter(self.enhanced_file_name, engine='openpyxl', mode='a') as writer:
        #     result_df.to_excel(writer, sheet_name=sheet_name, index=False)

        #     # Access the worksheet
        #     worksheet = writer.sheets[sheet_name]

        #     # Set uniform column width for all columns
        #     uniform_width = 10

        #     # Create header style
        #     header_fill = openpyxl.styles.PatternFill(
        #         start_color='003366',  # Dark blue
        #         end_color='003366',
        #         fill_type='solid'
        #     )
        #     header_font = openpyxl.styles.Font(
        #         color='FFFFFF',  # White
        #         bold=True
        #     )
        #     center_alignment = openpyxl.styles.Alignment(
        #         wrap_text=True,
        #         vertical='center',
        #         horizontal='center'
        #     )

        #     thin_border = openpyxl.styles.Border(
        #         left=openpyxl.styles.Side(style='thin'),
        #         right=openpyxl.styles.Side(style='thin'),
        #         top=openpyxl.styles.Side(style='thin'),
        #         bottom=openpyxl.styles.Side(style='thin')
        #     )

        #     # Apply formats
        #     for col_idx, col_name in enumerate(result_df.columns, 1):
        #         column_letter = openpyxl.utils.get_column_letter(col_idx)
                
        #         # Set uniform width for each column
        #         worksheet.column_dimensions[column_letter].width = uniform_width

        #         # Format all cells in column
        #         for row_idx, row in enumerate(
        #             worksheet.iter_rows(min_row=1, max_row=len(result_df)+1,
        #                                 min_col=col_idx, 
        #                                 max_col=col_idx),
        #                                 1):
        #             for cell in row:
        #                 if row_idx == 1:
        #                     cell.fill = header_fill
        #                     cell.font = header_font
                            
        #                 # Apply centering and borders to all cells
        #                 cell.alignment = center_alignment
        #                 cell.border = thin_border

        #     # Set header row height for the possibility of wrapped text
        #     worksheet.row_dimensions[1].height = 30

        # print(f"Pivot table added to {self.enhanced_file_name} in '{sheet_name}' sheet.")


    def insert_pivot_tables(self):
        '''Insert pivot tables based on the data in the original file'''
        if self.enhanced_file_name is None:
            raise ValueError("Enhanced file name is not set. Call add_context_columns() first.")
        
         # Load the enhanced Excel file
        try:
            df = pd.read_excel(self.enhanced_file_name)
        except FileNotFoundError:
            raise FileNotFoundError(f"The file {self.enhanced_file_name} does not exist.")
        except Exception as e:
            raise Exception(f"An error occurred while reading the file: {str(e)}")
        
        pairs = []
        sheet_list = []
        
        count_pairs, count_sheet_list = self.insert_count_pivot_tables(df)

        # result_df, sheet_name = self.calculate_count_pivot_by_column(df, "Date")
        # pairs.append((result_df, sheet_name))
        # sheet_list.append(sheet_name)

        # result_df, sheet_name = self.calculate_count_pivot_by_column(df, "Hour")
        # pairs.append((result_df, sheet_name))
        # sheet_list.append(sheet_name)

        # result_df, sheet_name = self.calculate_count_pivot_by_column(df, "Weekday")
        # pairs.append((result_df, sheet_name))
        # sheet_list.append(sheet_name)

        pairs = pairs + count_pairs
        sheet_list = sheet_list + count_sheet_list

        self.write_pivot_sheets(pairs)

        self.format_pivot_sheets(sheet_list)

    def insert_count_pivot_tables(self, df: pd.DataFrame):

        pairs = []
        sheet_list = []
        
        result_df, sheet_name = self.calculate_count_pivot_by_column(df, "Date")
        pairs.append((result_df, sheet_name))
        sheet_list.append(sheet_name)

        result_df, sheet_name = self.calculate_count_pivot_by_column(df, "Hour")
        pairs.append((result_df, sheet_name))
        sheet_list.append(sheet_name)

        result_df, sheet_name = self.calculate_count_pivot_by_column(df, "Weekday")
        pairs.append((result_df, sheet_name))
        sheet_list.append(sheet_name)

        return pairs, sheet_list



    def calculate_count_pivot_by_column(self, input_df: pd.DataFrame, pivot_column="Date"):
        '''This method calculated all the pivot tables that relate to measures
        that implement counts (such as 'request count').'''
         # Chech the required columns exist in the Excel file
        required_columns = ['Date', pivot_column]
        missing = [col for col in required_columns if col not in input_df.columns]
        if missing:
            raise ValueError(f"The enhanced file is missing required columns: {', '.join(missing)}")
        
        # include_time_keyboards = ["Time"]
        # exclude_time_keyboards = ["Timestamp", "Compliance"]
        # include_count_keywords = ["Count", "Total"]
        # include_rate_keword = ["Compliance", "Rate"]
        
        # Create result data frame to include all columns
        # If pivoting by weekday, make sure the original list is ordered
        if pivot_column == "Weekday":
            weekday_order = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            unique_weekdays = input_df[pivot_column].unique()
            ordered_weekdays = [day for day in weekday_order if day in unique_weekdays]
            result_df = pd.DataFrame({pivot_column: ordered_weekdays})
        else:
            result_df = pd.DataFrame({pivot_column: sorted(input_df[pivot_column].unique())})

        # Iterate through columns to find all the ones to include
        for col in input_df.columns:
            if "Count" in col or "Total" in col:

                # Maintain the original column name for clarity, no need for changes

                # Separate different cases: time_column is "Date", or something else
                if pivot_column == "Date":
                    dimension_df = input_df.groupby("Date", as_index=False)[col].sum()
                else:
                    # Two steps to summarize: step 1 - Sum by Date and time_column ("Hour", "Weekday", etc.)
                    daily_totals_df = input_df.groupby(["Date", pivot_column], as_index=False)[col].sum()

                    # Step 2 - Average across dates for each time_column
                    dimension_df = daily_totals_df.groupby(pivot_column, as_index=False)[col].mean()

                # Convert to int (as these are counts)
                dimension_df[col] = dimension_df[col].astype(int)

                # Merge with result data frame
                result_df = result_df.merge(
                    dimension_df,
                    on=pivot_column,
                    how="left"
                )
        
        # Define the name for the new sheet according to situation
        sheet_name = 'Count Pivot' if pivot_column == "Date" else f'Count Pivot - {pivot_column}'
        if len(sheet_name) > 31:  # Excel sheet names must be <= 31 characters
            sheet_name = sheet_name[:31]

        return result_df, sheet_name
    

    def write_pivot_sheets(self, pairs: List[Tuple[pd.DataFrame, str]]) -> None:
        '''This method writes all the previously calculated pivot tables.
        Take a list of data frames and sheet names to write in Excel'''
        with pd.ExcelWriter(self.enhanced_file_name, engine='openpyxl', mode='a') as writer:
            for df, sheet_name in pairs:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Pivot table added to {self.enhanced_file_name} in '{sheet_name}' sheet.")

    
    def format_pivot_sheets(self, sheet_list):
        '''Takes a list of the pivot sheets that have been created for a given file.
        Uses this list to add the formatting options required for this each sheet.'''

        print("Formatting all pivot sheets...")

        # Open Excel file
        wb = openpyxl.load_workbook(self.enhanced_file_name)
        
        # Go through all the sheets, one by one.
        for sheet_name in sheet_list:
            if sheet_name not in wb.sheetnames:
                continue

            print(f"Formatting sheet: {sheet_name}")

            # Access the worksheet
            worksheet = wb[sheet_name]

            # Set uniform column width for all columns
            uniform_width = 10

            # Create header style
            header_fill = openpyxl.styles.PatternFill(
                start_color='003366',  # Dark blue
                end_color='003366',
                fill_type='solid'
            )
            header_font = openpyxl.styles.Font(
                color='FFFFFF',  # White
                bold=True
            )
            center_alignment = openpyxl.styles.Alignment(
                wrap_text=True,
                vertical='center',
                horizontal='center'
            )

            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

            # Apply formats
            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                
                # Set uniform width for each column
                worksheet.column_dimensions[column_letter].width = uniform_width

                # Format all cells in column
                for row_idx in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)

                    
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                            
                    # Apply centering and borders to all cells
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # Set header row height for the possibility of wrapped text
            worksheet.row_dimensions[1].height = 30

        # Save changes to all sheets
        print(f"Saving all formatting changes to file {self.enhanced_file_name}...")
        wb.save(self.enhanced_file_name)


        


        


     


    def create_charts(self):
        """Create charts based on the pivot DataFrame."""

        # Open the enhanced Excel file, that already has the pivot table
        # to create charts based on the pivot table data.
        # Use libraries like openpyxl or xlsxwriter to create charts.
        wb = load_workbook(self.enhanced_file_name)

        # Verify if the 'Pivot' sheet exists
        if "Pivot" not in wb.sheetnames:
            raise ValueError("The 'Pivot' sheet does not exist in the enhanced file.")
        
        # Retrieve the pivot sheet
        pivot_sheet = wb["Pivot"]

        # rows = [
        #     ('Date', 'Average Time'),
        #     ('2023-10-01', 12.0),
        #     ('2023-10-02', 15.0),
        #     ('2023-10-03', 10.0),
        #     ('2023-10-04', 20.0),
        #     ('2023-10-05', 18.0),
        #     ('2023-10-06', 22.0),

        # ]

        # my_sheet = wb.create_sheet(title="Test Data")

        # for row in rows:
        #     my_sheet.append(row)


        # Create a new sheet for charts
        chart_sheet = wb.create_sheet(title="Charts")

        # Crete a bar chart
        chart = BarChart()
        chart.type = "col" # Column chart (vertical bars)
        chart.style = 1 # Style of the chart (built-in style)
        chart.x_axis.majorGridlines = None # No gridlines on the x-axis

        chart.title = "Average Response Time by Date"
        # chart.x_axis.title = "Date"
        # chart.y_axis.title = "Average Time (s)"
        # chart.legend = None # No legend

        # Reference data for the chart. The data is contained in a sheet called 'Pivot'.
        # In simple words, column A contains the dates and column B contains the average response times.
        # There is only one row with the column names, so we can use min_row=2 to start from the second row.
        # There is also only one data series, which contains the average response times by date.
        values = Reference(pivot_sheet, min_col=2, min_row=1, max_row=pivot_sheet.max_row)
        categories = Reference(pivot_sheet, min_col=1, min_row=1, max_row=pivot_sheet.max_row)

        # Add data to the chart
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
   
        


        # Verify we have exactly one series and so series formatting
        if len(chart.series) == 1:
            # Get the first (and only) series
            series = chart.series[0]
            
            # This is how you properly modify series properties
            series.tx = SeriesLabel(v="Average Time")
            
            # Ensure consistent formatting
            series.graphicalProperties.solidFill = "4472C4"  # Standard Excel blue


        else:
            raise ValueError("Chart should have exactly one series")

        # Add the chart to the Charts sheet at cell B2
        chart_sheet.add_chart(chart, "B2")

        # Save the charts in the enhanced file
        self.enhanced_file_name = self.enhanced_file_name.replace('.xlsx', '_with_charts.xlsx')
        wb.save(self.enhanced_file_name)
        print(f"Charts created and saved in {self.enhanced_file_name} in 'Charts' sheet.")


        # # Create bar chart
        # chart = BarChart()
        # chart.title = "Average Response Time by Date"
        # chart.x_axis.title = "Date"
        # chart.y_axis.title = pivot_sheet.cell(row=1, column=2).value

        # # Reference data
        # data = Reference(pivot_sheet, min_col=2, min_row=2, max_row=pivot_sheet.max_row)
        # cats = Reference(pivot_sheet, min_col=1, min_row=2, max_row=pivot_sheet.max_row)

        # chart.add_data(data, titles_from_data=True)
        # chart.set_categories(cats)
        
        # chart.legend = None

        # # Add chart to Charts sheet
        # chart_sheet.add_chart(chart, "B2")

        # self.enhanced_file_name = self.enhanced_file_name.replace('.xlsx', '_with_charts.xlsx')
        
        # # Save the workbook with the new chart
        # wb.save(self.enhanced_file_name)
        # print(f"Charts created and saved in {self.enhanced_file_name} in 'Charts' sheet.")


    def create_test_chart(self):
        """Create a test chart with dummy data."""
        # This method is for testing purposes only.
        # It creates a chart with dummy data to verify the chart creation functionality.
        
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()

        rows = [
            ('Number', 'Batch 1', 'Batch 2'),
            (2, 10, 30),
            (3, 40, 60),
            (4, 50, 70),
            (5, 20, 10),
            (6, 10, 40),
            (7, 50, 30),
        ]

        ws = wb.create_sheet()
        for row in rows:
            ws.append(row)

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Test Chart"
        chart1.x_axis.title = "Test Number"
        chart1.y_axis.title = "Test Value"

        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=len(rows))
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(rows))

        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4

        chart1.legend = None
        chart1.height = 20 
        chart1.width = 30

        chart1.title.tx.rich.p[0].pPr.defRPr = CharacterProperties(sz=2400, b=True)  # Set title font size and bold

        chart1.plot_area.barChart.series[0].cat.strRef.strCache.pt[0].v = "2"  # Set x-axis title
        




        # Format x-axis title font to be bigger and bold
        

        # chart1.x_axis.majorGridlines = ChartLines()
        # chart1.y_axis.majorGridlines = ChartLines() 

        ws.add_chart(chart1, "B10")

        wb.save("output_files/enhanced/test_chart.xlsx") 

        

def main():
    print("Hello from excel_enhancer.py!\n")
    print(os.getcwd())

    # Create ExcelEnhancer instance
    enhancer = ExcelEnhancer("output_files/ComercioTransaccionesController_DEFAULT.xlsx")

    # Add context columns to the Excel file
    try:
        enhancer.add_context_columns()
    except Exception as e:
        print(f"An error occurred: {str(e)}")

    enhancer.set_file_name("output_files/enhanced/ComercioTransaccionesController_DEFAULT_ENH.xlsx")

    # Insert pivot datsable based on time context columns
    enhancer.insert_time_pivot()

    # Insert pivot datsable based on compliance context columns
    enhancer.insert_compliance_pivot()

    # Insert pivot datsable based on count context columns
    # enhancer.insert_count_pivot()

    # Insert chart for the pivot table (in a separete sheet
    # enhancer.create_test_chart()
    enhancer.insert_pivot_tables()

    # Take enhanced file name and add "with_charts" suffix
    # excel_file = enhancer.enhanced_file_name

    # print(f"Source file for the creation of files: {excel_file}")

    # charts_dll = r"/Users/agustin/Documents/C#/ExcelChartTest/bin/Release/net9.0/ExcelChartTest.dll"

    # if not os.path.exists(charts_dll):
    #     raise FileNotFoundError(f"DLL not found at: {charts_dll}")
    # if not os.path.exists(excel_file):
    #     raise FileNotFoundError(f"Excel file not found at: {excel_file}")
    
    # dotnet_path = "/usr/local/share/dotnet/dotnet"

    # command = [dotnet_path, 
    #            "exec",
    #            charts_dll,
    #            excel_file
    #            ]

    # print(f"Running command: {' '.join(command)}")

    # try:
    #     result = subprocess.run(
    #         command,
    #         check=True,
    #         capture_output=True,
    #         text=True
    #         )

    #     print("Ran command successfully...")
    #     print("Output:", result.stdout)
    #     print(f"Charts created and saved in {excel_file} in 'Charts' sheet.")

    # except subprocess.CalledProcessError as e:
    #     print(f"Error running command: {e}")
    #     print("Output:", e.stdout)
    #     print("Error Output:", e.stderr)
    #     raise

    


    

    print("The end")



if __name__ == "__main__":
    main()